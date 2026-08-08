# -*- coding: utf-8 -*-
"""
多表合并透视工具（python-Excel）

功能：
1. 支持一次粘贴/拖入多个 Excel 表格（*.xlsx / *.xlsm），回车确认；
   - 资源管理器多选后 Ctrl+C，再到终端 Ctrl+V 多行粘贴（每行一个路径）；
   - 也可输入 c 从剪贴板直接读取全部路径（参考 pdf-v-photo/pdf转图片.py）。
2. 列出全部表名（合并去重）与全部列名（合并去重），选择要合并的列；
   子表列名所在行与工作表序号可配置（默认第 1 行、第 1 个工作表）。
3. 按粘贴顺序追加合并所选列，每行末尾新增“数据来源表”列。
4. 透视汇总：
   - 固定模式：按 运单号、币种 分组，对 运费、附加费1/2/3、报关费、合计金额 求和；
   - 自定义模式：自由选择行标签列与求和数值列。
5. 透视表去向：
   - 直接保存为新的 xlsx 文件；
   - 或插入粘贴的主表中作为新工作表，并按“运单号”回填透视表“合计金额”列，
     再新增一列“透视表合计金额 - 主表汇总金额”的差异列；
     主表列名从第 3 行开始（可输入调整），“汇总金额”一般位于 AH 列。

使用：
    python 多表合并透视.py
"""

import os
import re
import sys
from collections import OrderedDict
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# 控制台颜色
# ---------------------------------------------------------------------------
# ANSI 转义序列：交互提示用加粗亮绿色，便于与输出信息区分
COLOR_PROMPT = '\033[1;92m'  # 加粗亮绿色（提示文字）
COLOR_RESET = '\033[0m'      # 复位
# 仅当标准输出是终端时启用颜色，避免重定向/管道出现乱码
_USE_COLOR = sys.stdout.isatty()


def init_console_color():
    """Windows 控制台启用 ANSI 转义序列（VT 处理），其他平台无需处理。"""
    if _USE_COLOR and os.name == 'nt':
        try:
            import ctypes
            kernel32 = ctypes.windll.kernel32
            h = kernel32.GetStdHandle(-11)
            mode = ctypes.c_ulong()
            if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
                kernel32.SetConsoleMode(h, mode.value | 0x0004)  # ENABLE_VIRTUAL_TERMINAL_PROCESSING
        except Exception:
            pass


# ---------------------------------------------------------------------------
# 输入工具（参考 pdf-v-photo/pdf转图片.py）
# ---------------------------------------------------------------------------
# 匹配拖入路径：优先引号包裹（路径可含空格），否则连续非空白
PATH_RE = re.compile(r'"([^"]*)"|(\S+)')

# 预读缓冲：多行粘贴时，非路径行先暂存，让下一个问题优先读取
_PUTBACK = []


def _ask(prompt):
    """读一行；若预读缓冲有内容则优先取缓冲，否则阻塞等待输入。"""
    if _PUTBACK:
        return _PUTBACK.pop(0)
    if _USE_COLOR:
        prompt = COLOR_PROMPT + prompt + COLOR_RESET
    return input(prompt)


class GoBack(Exception):
    """用户输入了「返回上一步」"""


def _ask_back(prompt):
    """同 _ask，但支持输入 b / back / 上一步 / 返回 触发 GoBack 异常。"""
    ans = _ask(prompt).strip()
    if ans.lower() in ('b', 'back', '上一步', '返回'):
        raise GoBack()
    return ans


def is_path_like(s):
    """粗略判断一行是否为文件路径（存在该文件，或含路径分隔符/盘符）。"""
    if not s:
        return False
    if os.path.isfile(s):
        return True
    if '/' in s or '\\' in s or ':' in s:
        return True
    return False


def split_line_paths(line):
    """把一行拆成多个路径（引号包裹优先，否则连续非空白）。"""
    parts = []
    for m in PATH_RE.finditer(line):
        p = m.group(1) if m.group(1) is not None else m.group(2)
        p = p.strip()
        if p:
            parts.append(p)
    return parts


def read_paths(prompt):
    """读取一行输入，按 Windows 拖拽/粘贴形态拆分成多个路径。
    兼容多种形态：
    - 单行单个/多个：d:/a.xlsx、"d:/a.xlsx" "d:/b c.xlsx"、d:/a.xlsx d:/b.xlsx
    - 输入 c 时从剪贴板读取全部路径（资源管理器多选后 Ctrl+C）
    - Ctrl+V 多行粘贴：每行一个路径，自动连续读取；非路径行暂存给下一个问题
    """
    first = _ask(prompt).strip()
    if first.lower() in ('c', 'cb', 'clip', '粘贴'):
        return read_clipboard_paths()
    paths = []
    if not first:
        return []
    if not is_path_like(first):
        return []  # 非路径，交由上层重试
    if os.path.isfile(first):
        paths.append(first)
    else:
        paths.extend(split_line_paths(first))
    # 继续读取后续路径行（支持 Ctrl+V 多行粘贴）
    while True:
        nxt = _ask('').strip()
        if not nxt:
            break
        if is_path_like(nxt):
            if os.path.isfile(nxt):
                paths.append(nxt)
            else:
                paths.extend(split_line_paths(nxt))
        else:
            _PUTBACK.append(nxt)
            break
    return paths


def read_clipboard_paths():
    """从剪贴板读取文件路径列表（资源管理器多选后 Ctrl+C 时，每行一个引号路径）。"""
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        root.update()
        try:
            data = root.clipboard_get()
        except tk.TclError:
            return []
        finally:
            root.destroy()
        paths = []
        for line in data.splitlines():
            line = line.strip()
            if not line:
                continue
            paths.extend(split_line_paths(line))
        return paths
    except Exception:
        return []


def read_save_path(prompt, default_name):
    """读取保存路径：支持直接输入、拖入文件夹/文件（自动去掉 Windows 拖入时带的引号）。
    空输入用默认名；拖入文件夹则在其下生成默认文件名；无扩展名补 .xlsx。"""
    s = _ask(prompt).strip()
    if len(s) >= 2 and s[0] == s[-1] == '"':
        s = s[1:-1].strip()
    if not s:
        return default_name
    if os.path.isdir(s):
        return os.path.join(s, default_name)
    if not s.lower().endswith(('.xlsx', '.xlsm')):
        s += '.xlsx'
    return s


# ---------------------------------------------------------------------------
# 通用工具
# ---------------------------------------------------------------------------
def parse_multi_choice(text, n):
    """把 '1,3,5-7' 解析为 0 起索引列表（去重、校验范围）；支持 all/全部/*。"""
    t = text.strip()
    if not t:
        return []
    if t.lower() in ('all', '全部', '*'):
        return list(range(n))
    idxs = []
    for part in re.split(r'[,\s，、]+', t):
        if not part:
            continue
        if '-' in part or '~' in part:
            a, b = re.split(r'[-~]', part, maxsplit=1)
            try:
                a, b = int(a), int(b)
            except ValueError:
                raise ValueError('无效范围：%s' % part)
            if a > b:
                a, b = b, a
            if a < 1 or b > n:
                raise ValueError('序号超出范围：%s' % part)
            idxs.extend(range(a - 1, b))
        else:
            try:
                i = int(part)
            except ValueError:
                raise ValueError('无效序号：%s' % part)
            if i < 1 or i > n:
                raise ValueError('序号超出范围：%s' % part)
            idxs.append(i - 1)
    seen = set()
    out = []
    for i in idxs:
        if i not in seen:
            seen.add(i)
            out.append(i)
    return out


def sanitize_sheet_name(name):
    """去掉工作表名中的非法字符。"""
    name = re.sub(r'[\\/:*?"<>|\r\n]', '_', str(name)).strip()
    return name or '透视汇总'


def to_number(v):
    """把单元格值转成可求和的数值，无法解析的按 0 处理。"""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('，', '')
    s = re.sub(r'[¥￥$€\s]', '', s)
    s = s.rstrip('元')
    if s in ('', '-', '--', '/', 'N/A', 'NA'):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_key(v):
    """把值规范成用于匹配的文本键（运单号匹配用，兼容数字/字符串）。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return repr(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def find_column(headers, candidates):
    """在表头列表中找列，返回 0 起列索引或 None；优先精确匹配，再匹配包含。"""
    cands = [c for c in candidates if c]
    for c in cands:
        if c in headers:
            return headers.index(c)
    for i, h in enumerate(headers):
        hs = str(h).strip()
        for c in cands:
            if c in hs:
                return i
    return None


def find_col_in_table(headers, name, aliases):
    """在子表表头中找列：精确匹配所选名 → 精确匹配替代名 → 模糊包含匹配。
    返回 0 起列索引或 None。"""
    h = {str(x).strip(): i for i, x in enumerate(headers)}
    for cand in [name] + list(aliases):
        if cand in h:
            return h[cand]
    for cand in [name] + list(aliases):
        for i, x in enumerate(headers):
            xs = str(x).strip()
            if cand in xs:
                return i
    return None


def choose_column(prompt, headers, default_candidates=None, default_label=None):
    """列出列名供用户选择，返回选中的列名（原样）。
    headers：列名列表（可能含 None/空）；default_candidates：默认推荐列名（取第一个命中的）。
    支持输入 b / 上一步 返回上一步。"""
    print('列名：')
    for i, h in enumerate(headers, start=1):
        hs = '' if h is None else str(h).strip()
        print('  %d. %s' % (i, hs if hs else '（空）'))
    default_idx = 1
    if default_candidates:
        flat = ['' if h is None else str(h).strip() for h in headers]
        for cand in default_candidates:
            idx = find_column(flat, [cand])
            if idx is not None:
                default_idx = idx + 1
                break
    label = default_label or ('回车默认 %d' % default_idx)
    while True:
        sel = _ask_back('%s（输入序号，%s）：' % (prompt, label)).strip()
        if not sel:
            return headers[default_idx - 1]
        if sel.isdigit() and 1 <= int(sel) <= len(headers):
            return headers[int(sel) - 1]
        print('请输入有效序号（1-%d）。' % len(headers))


# ---------------------------------------------------------------------------
# 读取表格
# ---------------------------------------------------------------------------
def read_table(path, sheet_name=None, sheet_idx=0, header_row=1):
    """读取 Excel 的指定工作表，返回 (工作表名, 表头列表, 数据行列表)。
    sheet_name 优先：按工作表名称取（所有子表统一用选中的工作表合并）；
    未指定时按 sheet_idx（0 起，默认第 1 个）取第几个工作表。
    header_row 为列名所在行（1 起，默认第 1 行），数据从 header_row+1 行开始。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ValueError('该文件没有工作表"%s"' % sheet_name)
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[sheet_idx]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if len(rows) < header_row:
        return ws.title, [], []
    header_cells = rows[header_row - 1]
    headers = []
    for i, v in enumerate(header_cells):
        name = v if isinstance(v, str) else ('' if v is None else str(v))
        name = name.strip()
        if name == '':
            name = '未命名列%d' % (i + 1)
        headers.append(name)
    data = [r for r in rows[header_row:] if any(v is not None and str(v).strip() != '' for v in r)]
    return ws.title, headers, data


# ---------------------------------------------------------------------------
# 透视汇总
# ---------------------------------------------------------------------------
def make_pivot(headers, rows, row_labels, value_cols, add_total=False):
    """按行标签分组，对数值列求和，返回 (透视表头, 透视表行)。
    add_total 为 True 时，在末尾追加「合计」列 = 各求和列相加。"""
    idx = {name: i for i, name in enumerate(headers)}
    missing = [c for c in row_labels + value_cols if c not in idx]
    if missing:
        raise KeyError('汇总表中不存在列：%s' % '、'.join(missing))
    rl_idx = [idx[x] for x in row_labels]
    v_idx = [idx[x] for x in value_cols]
    groups = OrderedDict()
    for r in rows:
        key = tuple(r[i] if i < len(r) else None for i in rl_idx)
        if all(norm_key(k) == '' for k in key):
            continue  # 行标签全空的行跳过
        if key not in groups:
            groups[key] = [to_number(r[i] if i < len(r) else None) for i in v_idx]
        else:
            sums = groups[key]
            for j, i in enumerate(v_idx):
                sums[j] += to_number(r[i] if i < len(r) else None)
    pivot_headers = list(row_labels) + ['求和项:%s' % c for c in value_cols]
    if add_total:
        pivot_headers.append('合计')
    pivot_rows = []
    for key, sums in groups.items():
        row = list(key) + sums
        if add_total:
            row.append(sum(sums))
        pivot_rows.append(row)
    return pivot_headers, pivot_rows


# ---------------------------------------------------------------------------
# 写表
# ---------------------------------------------------------------------------
def write_table(path, sheet_title, headers, rows):
    """把表头+数据行写入新的 xlsx 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 回填透视合计金额 + 差异列
# ---------------------------------------------------------------------------
def backfill_pivot_total(wb, sheet_name, pivot_headers, pivot_rows, header_row=3,
                         p_val_name=None, main_val_name=None):
    """在指定工作表中按运单号回填透视表的指定金额列，并新增差异列。
    主表列名位于 header_row 行，数据从 header_row+1 行开始；
    p_val_name：透视表中要回填的金额列名；
    main_val_name：主表中要对比的金额列名；
    返回 (差异行数, 透视金额列名, 差异列名)。"""
    ws = wb[sheet_name]
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    headers_str = ['' if v is None else str(v).strip() for v in headers]

    p_headers_str = ['' if v is None else str(v).strip() for v in pivot_headers]
    p_yd_idx = find_column(p_headers_str, ['运单号'])
    if p_yd_idx is None:
        raise ValueError('透视表中未找到“运单号”列，无法回填。')
    p_hz_idx = find_column(p_headers_str, [p_val_name]) if p_val_name else None
    if p_hz_idx is None:
        raise ValueError('透视表中未找到金额列“%s”，无法回填。' % p_val_name)

    # 透视表按运单号建索引（重复取第一个）
    pivot_map = {}
    dup = 0
    for row in pivot_rows:
        k = norm_key(row[p_yd_idx] if p_yd_idx < len(row) else None)
        if not k:
            continue
        if k in pivot_map:
            dup += 1
            continue
        pivot_map[k] = row[p_hz_idx] if p_hz_idx < len(row) else None
    if dup:
        print('提示：透视表中有 %d 个重复运单号（如不同币种），回填时使用第一个匹配值。' % dup)

    # 新列名：透视列名去掉「求和项:」前缀
    base = str(p_val_name).replace('求和项:', '')
    p_display = '透视表%s' % base
    d_display = '%s差异（透视-原表）' % base

    # 新列插入位置：主表「备注」列右边；找不到备注则追加末尾
    bz_idx = find_column(headers_str, ['备注'])
    if bz_idx is not None:
        pos = bz_idx + 2  # 1-based 备注列号 + 1
    else:
        pos = ws.max_column + 1

    pf_exist = find_column(headers_str, [p_display])
    diff_exist = find_column(headers_str, [d_display])
    if pf_exist is not None and diff_exist is not None:
        # 已有这两列：先删除（列号大的先删，避免移位错乱），再重新插到备注右边
        ws.delete_cols(max(pf_exist, diff_exist) + 1, 1)
        ws.delete_cols(min(pf_exist, diff_exist) + 1, 1)
    ws.insert_cols(pos, amount=2)
    col_pf, col_diff = pos, pos + 1
    ws.cell(row=header_row, column=col_pf, value=p_display)
    ws.cell(row=header_row, column=col_diff, value=d_display)

    # 插入列后重新读表头，定位 运单号 与 对比列
    headers2 = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    headers_str2 = ['' if v is None else str(v).strip() for v in headers2]
    yd_idx = find_column(headers_str2, ['运单号'])
    if yd_idx is None:
        raise ValueError('所选工作表未找到“运单号”列，无法回填。')
    hz_idx = find_column(headers_str2, [main_val_name]) if main_val_name else None
    if hz_idx is None:
        raise ValueError('所选工作表未找到主表对比列“%s”，无法计算差异。' % main_val_name)

    filled = 0
    for r in range(header_row + 1, ws.max_row + 1):
        key = norm_key(ws.cell(row=r, column=yd_idx + 1).value)
        pf = pivot_map.get(key)
        ws.cell(row=r, column=col_pf, value=pf)
        if pf is None or norm_key(pf) == '':
            continue  # 透视表无此运单号，差异留空
        orig = ws.cell(row=r, column=hz_idx + 1).value
        if orig is None or norm_key(orig) == '':
            continue
        ws.cell(row=r, column=col_diff, value=to_number(pf) - to_number(orig))
        filled += 1
    return filled, p_display, d_display


# 主流程
# ---------------------------------------------------------------------------
def main():
    init_console_color()
    if not OPENPYXL_OK:
        print('提示：未安装 openpyxl，无法读写 Excel。请运行：pip install openpyxl')
        return

    print('多表合并透视工具')
    print('功能：粘贴多表 -> 选择列 -> 合并 -> 透视汇总 -> 保存或插入主表')
    print('提示：任意选择步骤中输入 b / 上一步 返回上一个步骤')
    print('-' * 64)

    step = 0  # 0=粘贴, 1=选工作表, 2=列名统一, 3=选列, 4=列名替代, 5=合并+保存, 6=透视, 7=去向
    valid = None; sheet_name = None; tables = None
    sel_cols = None; col_aliases = None
    merged_headers = None; merged_rows = None
    pivot_headers = None; pivot_rows = None

    while True:
        try:
            # ====== Step 0: 粘贴/读取表格 ======
            if step <= 0:
                while True:
                    paths = read_paths('请粘贴要合并的表格文件（可多选复制后 Ctrl+V 粘贴，或输入 c 从剪贴板读取），粘贴后回车：')
                    if not paths:
                        print('未检测到文件路径，请重新粘贴。')
                        continue
                    valid, seen = [], set()
                    for p in paths:
                        ap = os.path.normcase(os.path.abspath(p))
                        if ap in seen:
                            continue
                        seen.add(ap)
                        if not os.path.isfile(p):
                            print('跳过（找不到文件）：%s' % p)
                            continue
                        if not (p.lower().endswith('.xlsx') or p.lower().endswith('.xlsm')):
                            print('跳过（仅支持 .xlsx/.xlsm）：%s' % p)
                            continue
                        valid.append(p)
                    if not valid:
                        print('没有可用的表格文件，请重新粘贴。')
                        continue
                    break

                if len(valid) == 1:
                    print('已识别 1 个表格：%s' % os.path.basename(valid[0]))
                else:
                    print('已识别 %d 个表格（顺序即合并顺序）：' % len(valid))
                    for i, p in enumerate(valid, start=1):
                        print('  %d. %s' % (i, os.path.basename(p)))
                step = 1

            # ====== Step 1: 选择工作表 + 读取数据 ======
            if step <= 1:
                # ---- 1.5 选择要合并的工作表 ----
                all_sheets, seen_sheets = [], set()
                for p in valid:
                    try:
                        wb = load_workbook(p, read_only=True)
                        sheets = wb.sheetnames
                        wb.close()
                    except Exception as e:
                        print('跳过（无法读取工作表列表）：%s（%s）' % (os.path.basename(p), e))
                        continue
                    for s in sheets:
                        s = s.strip()
                        if s and s not in seen_sheets:
                            seen_sheets.add(s)
                            all_sheets.append(s)
                if not all_sheets:
                    print('无法读取任何工作表名称，请检查表格文件。')
                    return
                print('全部工作表名称（合并去重）：')
                for i, s in enumerate(all_sheets, start=1):
                    print('  %d. %s' % (i, s))
                while True:
                    sel = _ask_back('请选择要合并的工作表（输入序号或工作表名称，回车默认 1）：').strip()
                    if not sel:
                        sheet_name = all_sheets[0]
                        break
                    if sel.isdigit():
                        idx = int(sel)
                        if 1 <= idx <= len(all_sheets):
                            sheet_name = all_sheets[idx - 1]
                            break
                        print('序号超出范围（1-%d）。' % len(all_sheets))
                        continue
                    if sel in seen_sheets:
                        sheet_name = sel
                        break
                    print('不存在工作表名称：%s' % sel)
                print('已选择工作表：%s（全部表格都使用该工作表的数据合并）' % sheet_name)

                # 列名默认在第 1 行读取
                header_row = 1

                tables = []
                for p in valid:
                    try:
                        sheet_name, headers, rows = read_table(p, sheet_name=sheet_name, header_row=header_row)
                    except Exception as e:
                        print('读取失败，跳过：%s（%s）' % (os.path.basename(p), e))
                        continue
                    tables.append({'path': p, 'base': os.path.basename(p),
                                   'sheet': sheet_name, 'headers': headers, 'rows': rows})
                if not tables:
                    print('没有成功读取任何表格。')
                    return

                # 显示名：同名不同目录时附加目录
                name_counts = {}
                for t in tables:
                    name_counts[t['base']] = name_counts.get(t['base'], 0) + 1
                for t in tables:
                    if name_counts[t['base']] > 1:
                        t['display'] = '%s（%s）' % (t['base'], os.path.dirname(os.path.abspath(t['path'])))
                    else:
                        t['display'] = t['base']

                print('共读取 %d 个表格（顺序即合并顺序）：' % len(tables))
                for i, t in enumerate(tables, start=1):
                    print('  %d. %s（工作表：%s，%d 行数据）' % (i, t['display'], t['sheet'], len(t['rows'])))
                step = 2

            # ====== Step 2: 列名统一 ======
            if step <= 2:
                # ---- 1.6 列名统一：检测各表列名是否一致，可改名对齐 ----
                for i, t in enumerate(tables):
                    m = re.search(r'(\d+月)', t['base'])
                    t['short'] = m.group(1) if m else '表%d' % (i + 1)

                col_owners = {}
                for t in tables:
                    for h in t['headers']:
                        col_owners.setdefault(h, []).append(t['short'])
                common_cols = [c for c, o in col_owners.items() if len(o) == len(tables)]
                partial_cols = {c: o for c, o in col_owners.items() if len(o) < len(tables)}

                print('检测各表列名一致性：')
                if common_cols:
                    shown = '、'.join(common_cols[:6])
                    if len(common_cols) > 6:
                        shown += '…（共%d个）' % len(common_cols)
                    print('  一致（所有表都有）：%s' % shown)
                if partial_cols:
                    print('  不一致（部分表才有）：')
                    for c, owners in partial_cols.items():
                        missing = [t['short'] for t in tables if c not in t['headers']]
                        print('    「%s」只有 %s 有，缺 %s' % (c, '、'.join(owners), '、'.join(missing)))
                else:
                    print('  各表列名完全一致。')

                unify = False
                if partial_cols:
                    choice = _ask_back('检测到列名不一致。是否进入列名统一？\n'
                                       '  1 进入列名统一 | 2 跳过（后面选列时可用改名）。回车默认 1：').strip()
                    unify = (choice != '2')
                if unify:
                    while True:
                        print('各表列名：')
                        for i, t in enumerate(tables, start=1):
                            print('  %d.%s（%d列）：%s' % (i, t['short'], len(t['headers']), '、'.join(t['headers'])))
                        pick = _ask_back('请选择要改名的表（支持多个：2 3 4 5 或 2-5；输入 done 结束）：').strip()
                        if not pick or pick.lower() == 'done':
                            break
                        try:
                            t_idx = parse_multi_choice(pick, len(tables))
                        except ValueError as e:
                            print('选择无效：%s（支持 2 3 4 5 或 2-5）' % e)
                            continue
                        if not t_idx:
                            print('至少选一个表。')
                            continue

                        # 选一个表：先列该表列名再选列；选多个表：新名=旧名批量改
                        renames = []
                        if len(t_idx) == 1:
                            ti = t_idx[0]
                            print('%s 的列：' % tables[ti]['short'])
                            for j, h in enumerate(tables[ti]['headers'], start=1):
                                print('  %d. %s' % (j, h))
                            col_pick = _ask_back('请选择要改名的列（输入序号或列名；回车取消）：').strip()
                            if not col_pick:
                                print('已取消。')
                                continue
                            if col_pick.isdigit():
                                ci = int(col_pick) - 1
                                if not (0 <= ci < len(tables[ti]['headers'])):
                                    print('序号超出范围（1-%d）。' % len(tables[ti]['headers']))
                                    continue
                                old_name = tables[ti]['headers'][ci]
                            elif col_pick in tables[ti]['headers']:
                                old_name = col_pick
                            else:
                                print('该表没有此列：%s' % col_pick)
                                continue
                            new_name = _ask_back('「%s」改成什么列名？（回车取消）：' % old_name).strip()
                            if not new_name:
                                print('已取消。')
                                continue
                            renames.append((new_name, old_name))
                        else:
                            pair_in = _ask_back('批量改名（新名=旧名，多个用空格/逗号分隔，如：运单号=订单号 国家=Country；回车取消）：').strip()
                            if not pair_in:
                                print('已取消。')
                                continue
                            bad = False
                            for tok in re.split(r'[,\s，、]+', pair_in):
                                tok = tok.strip()
                                if not tok:
                                    continue
                                if '=' in tok:
                                    left, right = tok.split('=', 1)
                                    left, right = left.strip(), right.strip()
                                    if left and right:
                                        renames.append((left, right))
                                    else:
                                        print('格式不对：%s（应为 新名=旧名）' % tok)
                                        bad = True
                                else:
                                    print('格式不对：%s（应为 新名=旧名）' % tok)
                                    bad = True
                            if bad or not renames:
                                continue

                        done_names = []
                        for ti in t_idx:
                            hdrs = tables[ti]['headers']
                            for new_name, old_name in renames:
                                if old_name in hdrs:
                                    hdrs[hdrs.index(old_name)] = new_name
                                    done_names.append(tables[ti]['short'])
                        if done_names:
                            print('已改名：%s' % '、'.join(done_names))
                            for new_name, old_name in renames:
                                print('  「%s」→「%s」' % (old_name, new_name))
                        else:
                            print('所选表中没有这些列，未改名。')
                        col_owners = {}
                        for t in tables:
                            for h in t['headers']:
                                col_owners.setdefault(h, []).append(t['short'])
                        partial_cols = {c: o for c, o in col_owners.items() if len(o) < len(tables)}
                        if not partial_cols:
                            print('重新检测：各表列名已全部一致！')
                        else:
                            print('重新检测：仍不一致的列：%s' % '、'.join(partial_cols.keys()))
                        if _ask_back('是否继续改名？\n'
                                     '  1 继续改名 | 2 下一步（进入选列）。回车默认 1：').strip() == '2':
                            break
                step = 3

            # ====== Step 3: 选择要合并的列 ======
            if step <= 3:
                # ---- 2. 列选择 ----
                all_cols = []
                seen_cols = set()
                for t in tables:
                    for h in t['headers']:
                        if h not in seen_cols:
                            seen_cols.add(h)
                            all_cols.append(h)
                if not all_cols:
                    print('表格中没有可用的列名，无法合并。')
                    return
                while True:
                    print('全部列名（合并去重）：')
                    for i, c in enumerate(all_cols, start=1):
                        print('  %d. %s' % (i, c))
                    ans = _ask_back('请选择要合并的列（多个用逗号/空格分隔，如 1,3,5-7；输入 all 全选）：')
                    try:
                        sel_idx = parse_multi_choice(ans, len(all_cols))
                    except ValueError as e:
                        print('选择无效：%s，请重新输入。' % e)
                        continue
                    if not sel_idx:
                        print('至少选择一列。')
                        continue
                    sel_cols = [all_cols[i] for i in sel_idx]

                    missing_tables = {}
                    for t in tables:
                        hidx = {h: i for i, h in enumerate(t['headers'])}
                        miss = [c for c in sel_cols if c not in hidx]
                        if miss:
                            missing_tables[t['display']] = miss
                    if missing_tables:
                        for disp, miss in missing_tables.items():
                            print('提示：表 %s 缺少列 %s，这些列在该表留空。' % (disp, '、'.join(miss)))
                        if _ask_back('检测到有表格缺少所选列，是否重新选择列？(y/n，回车默认 y)：').strip().lower() in ('', 'y', 'yes'):
                            continue
                    break
                print('已选择 %d 列：%s' % (len(sel_cols), '、'.join(sel_cols)))
                step = 4

            # ====== Step 4: 列名替代映射（部分子表列名不同） ======
            if step <= 4:
                # 列名替代映射
                col_aliases = {}
                rename_ans = _ask_back('部分子表里的列名可能和汇总表列名不一致？\n'
                                      '  输入列名或序号，指定哪个列在子表里有不同的名字（回车默认跳过）：').strip()
                first_pick = None
                if rename_ans.lower() in ('', 'n', 'no', '否'):
                    pass
                elif rename_ans.isdigit():
                    idx = int(rename_ans)
                    if 1 <= idx <= len(sel_cols):
                        first_pick = sel_cols[idx - 1]
                    else:
                        print('序号超出范围（1-%d），进入改名流程。' % len(sel_cols))
                elif rename_ans.lower() in ('y', 'yes', '是'):
                    pass
                elif rename_ans in sel_cols:
                    first_pick = rename_ans
                else:
                    print('未识别输入（%s），进入改名流程。' % rename_ans)

                if first_pick is not None or rename_ans.lower() not in ('', 'n', 'no', '否'):
                    while True:
                        if first_pick is not None:
                            sel_name = first_pick
                            first_pick = None
                        else:
                            print('当前所选列：')
                            for i, c in enumerate(sel_cols, start=1):
                                print('  %d. %s' % (i, c))
                            sel_name = _ask_back('请选择在子表里叫法不同的列（输入序号或列名；输入 done 结束）：').strip()
                        if not sel_name or sel_name.lower() == 'done':
                            break
                        if sel_name.isdigit():
                            idx = int(sel_name)
                            if 1 <= idx <= len(sel_cols):
                                col = sel_cols[idx - 1]
                            else:
                                print('序号超出范围（1-%d）。' % len(sel_cols))
                                continue
                        elif sel_name in sel_cols:
                            col = sel_name
                        else:
                            print('所选列中不存在：%s' % sel_name)
                            continue
                        aliases = _ask_back('请输入「%s」在子表里的其他叫法（多个用逗号/空格分隔）：' % col).strip()
                        alias_list = [a.strip() for a in re.split(r'[,\s，、]+', aliases) if a.strip()]
                        if not alias_list:
                            print('未输入有效列名，跳过。')
                            continue
                        col_aliases[col] = alias_list
                        print('已设置：%s -> %s' % (col, '、'.join(alias_list)))
                    if col_aliases:
                        print('列名替代映射：' + '；'.join('%s -> %s' % (k, '、'.join(v)) for k, v in col_aliases.items()))
                step = 5

            # ====== Step 5: 追加合并 + 保存 ======
            if step <= 5:
                # ---- 3. 追加合并 ----
                merged_headers = list(sel_cols) + ['数据来源表']
                merged_rows = []
                for t in tables:
                    hidx_map = {}
                    missing = []
                    for c in sel_cols:
                        i = find_col_in_table(t['headers'], c, col_aliases.get(c, []))
                        hidx_map[c] = i
                        if i is None:
                            missing.append(c)
                    if missing:
                        print('提示：表 %s 缺少列 %s，这些列在该表留空。' % (t['display'], '、'.join(missing)))
                    for r in t['rows']:
                        row = []
                        for c in sel_cols:
                            i = hidx_map.get(c)
                            row.append(r[i] if i is not None and i < len(r) else None)
                        row.append(t['display'])
                        merged_rows.append(row)
                print('合并完成：共 %d 行，%d 列（含“数据来源表”）。' % (len(merged_rows), len(merged_headers)))

                # 可选：保存合并汇总表
                if _ask_back('是否保存合并后的汇总表？(y/n，回车默认 n)：').strip().lower() in ('y', 'yes'):
                    out = read_save_path('请输入保存路径（可拖入文件夹；回车默认：合并汇总.xlsx）：', '合并汇总.xlsx')
                    try:
                        write_table(out, '合并汇总', merged_headers, merged_rows)
                        print('已保存合并汇总表：%s' % out)
                    except Exception as e:
                        print('保存合并汇总表失败：%s' % e)
                step = 6

            # ====== Step 6: 透视 ======
            if step <= 6:
                # ---- 4. 透视模式 ----
                print('-' * 64)
                pivot_cols = [c for c in merged_headers if c != '数据来源表']
                while True:
                    print('合并后汇总表列（分组列自动为运单号、币种等非求和列）：')
                    for i, c in enumerate(pivot_cols, start=1):
                        print('  %d. %s' % (i, c))
                    ans = _ask_back('请选择要求和的数值列（金额列，多个用逗号/空格分隔，如 3,4,5,6,7）：')
                    try:
                        v_idx = parse_multi_choice(ans, len(pivot_cols))
                    except ValueError as e:
                        print('选择无效：%s，请重新输入。' % e)
                        continue
                    if not v_idx:
                        print('至少选一列求和。')
                        continue
                    break
                value_cols = [pivot_cols[i] for i in v_idx]
                # 自动分组：运单号、币种（存在则用）；否则用非求和列
                row_labels = [c for c in ['运单号', '币种'] if c in pivot_cols and c not in value_cols]
                if not row_labels:
                    row_labels = [c for i, c in enumerate(pivot_cols) if i not in v_idx]
                if not row_labels:
                    row_labels = [c for c in pivot_cols if c not in value_cols][:1] or [pivot_cols[0]]
                # 自动追加「合计」列 = 所选求和列之和（若未选含汇总/合计的列）
                add_total = not any(('合计' in c or '汇总' in c) for c in value_cols)
                print('分组列（自动）：%s' % '、'.join(row_labels))
                if add_total:
                    print('提示：透视表会自动生成「合计」列 = 所选求和列相加。')

                try:
                    pivot_headers, pivot_rows = make_pivot(pivot_cols, merged_rows, row_labels, value_cols, add_total=add_total)
                except KeyError as e:
                    print('透视失败：%s' % e)
                    return
                print('透视完成：共 %d 行 × %d 列。' % (len(pivot_rows), len(pivot_headers)))
                print('透视表列：%s' % '、'.join(pivot_headers))
                for i, r in enumerate(pivot_rows[:5], start=1):
                    print('  %d. %s' % (i, ' | '.join('' if v is None else str(v) for v in r)))
                if len(pivot_rows) > 5:
                    print('  ...（共 %d 行）' % len(pivot_rows))
                step = 7

            # ====== Step 7: 透视表去向 ======
            if step <= 7:
                # ---- 5. 透视表去向 ----
                print('-' * 64)
                while True:
                    dest = _ask_back('请选择透视表去向：1 直接保存为文件 | 2 插入到已有的主表（Excel）中。输入 1 或 2：').strip()
                    if dest == '1':
                        out = read_save_path('请输入保存路径（可拖入文件夹；回车默认当前目录 透视汇总.xlsx）：', '透视汇总.xlsx')
                        if os.path.exists(out):
                            base, ext = os.path.splitext(out)
                            out = '%s_%s%s' % (base, datetime.now().strftime('%Y%m%d_%H%M%S'), ext)
                        try:
                            write_table(out, '透视汇总', pivot_headers, pivot_rows)
                            print('已保存透视表：%s' % out)
                        except Exception as e:
                            print('保存透视表失败：%s' % e)
                        break
                    elif dest == '2':
                        while True:
                            mains = read_paths('请粘贴要插入的主表（*.xlsx），粘贴后回车：')
                            if not mains:
                                print('未检测到文件路径，请重新粘贴。')
                                continue
                            main_path = mains[0]
                            if not os.path.isfile(main_path):
                                print('找不到文件：%s' % main_path)
                                continue
                            if not (main_path.lower().endswith('.xlsx') or main_path.lower().endswith('.xlsm')):
                                print('仅支持 .xlsx/.xlsm 主表，请重新粘贴。')
                                continue
                            break
                        try:
                            wb = load_workbook(main_path)
                        except Exception as e:
                            print('打开主表失败：%s' % e)
                            return

                        sheet_name = sanitize_sheet_name(
                            _ask_back('请输入透视表工作表名称（回车默认：透视汇总）：').strip() or '透视汇总')
                        if sheet_name in wb.sheetnames:
                            del wb[sheet_name]
                        pws = wb.create_sheet(title=sheet_name)
                        pws.append(pivot_headers)
                        for row in pivot_rows:
                            pws.append(row)
                        print('已插入透视表工作表：%s' % sheet_name)

                        print('主表 %s 的全部工作表：' % os.path.basename(main_path))
                        for i, name in enumerate(wb.sheetnames, start=1):
                            print('  %d. %s' % (i, name))
                        while True:
                            sel = _ask_back('请选择要回填合计金额的工作表（输入序号，回车默认 1）：').strip() or '1'
                            try:
                                target = wb.sheetnames[int(sel) - 1]
                                break
                            except (ValueError, IndexError):
                                print('请输入有效序号（1-%d）。' % len(wb.sheetnames))
                        print('已选择工作表：%s' % target)

                        # 自动识别主表列名所在行（找含「运单号」的行；前 1-3 行多为合并标题），找不到再询问
                        target_ws = wb[target]
                        header_row = None
                        for rr in range(1, min(target_ws.max_row, 20) + 1):
                            row_str = ['' if target_ws.cell(rr, c).value is None
                                       else str(target_ws.cell(rr, c).value).strip()
                                       for c in range(1, min(target_ws.max_column, 80) + 1)]
                            if any('运单号' in s for s in row_str):
                                header_row = rr
                                break
                        if header_row is None:
                            while True:
                                hdr_in = _ask_back('请输入主表列名所在行（未自动识别到含「运单号」的行）：').strip()
                                if hdr_in.isdigit() and int(hdr_in) >= 1:
                                    header_row = int(hdr_in)
                                    break
                                print('请输入有效的行号（正整数）。')
                        else:
                            print('自动识别：主表列名在第 %d 行，数据从第 %d 行开始查找。' % (header_row, header_row + 1))
                        # 选择透视表要回填的金额列
                        p_val_name = choose_column(
                            '请选择透视表中要回填到主表的金额列',
                            pivot_headers,
                            default_candidates=['求和项:汇总', '求和项:合计金额', '汇总', '合计金额'],
                            default_label='回车默认推荐列')
                        # 选择主表中要对比的金额列
                        main_headers = [wb[target].cell(row=header_row, column=c).value
                                        for c in range(1, wb[target].max_column + 1)]
                        main_val_name = choose_column(
                            '请选择主表中要对比的金额列',
                            main_headers,
                            default_candidates=['汇总金额', '合计金额', '费用合计'],
                            default_label='回车默认推荐列')
                        try:
                            filled, f_col, d_col = backfill_pivot_total(
                                wb, target, pivot_headers, pivot_rows, header_row,
                                p_val_name=p_val_name, main_val_name=main_val_name)
                        except Exception as e:
                            print('回填失败：%s' % e)
                            return
                        wb.save(main_path)
                        print('已新增列：%s、%s，共计算 %d 行金额差异。' % (f_col, d_col, filled))
                        print('已保存主表：%s' % main_path)
                        break
                    else:
                        print('请输入 1 或 2。')
                step = 8

            # 全部完成
            break

        except GoBack:
            step -= 1
            if step < 0:
                step = 0
            print('已返回上一步（步骤 %d）。' % (step + 1))

    print('已完成，谢谢使用。')
    print('玛卡巴卡""')


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('\n用户取消操作。')
    sys.exit(0)