# -*- coding: utf-8 -*-
"""FastAPI 入口：任务接口 + 托管前端构建产物。

- POST /api/tasks                 上传 PDF（多选）并创建后台任务
- GET  /api/tasks/{id}            轮询任务状态/进度
- GET  /api/tasks/{id}/download   下载生成的 Excel
- GET  /                          前端页面（frontend/dist 存在时）
"""
import os
import sys
import tempfile
import uuid

from fastapi import FastAPI, File, Form, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import tasks  # noqa: E402

app = FastAPI(title='财务内部在线工具网页版', docs_url='/api/docs', openapi_url='/api/openapi.json')

_FRONT_DIST = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)), '..', 'frontend', 'dist'))
_XLSX_MEDIA = ('application/vnd.openxmlformats-officedocument.'
               'spreadsheetml.sheet')


@app.post('/api/worksheets')
async def read_worksheets(template: UploadFile = File(...)):
    """读取上传表格的工作表列表，供前端选工作表用。返回 {sheets: [...]}。"""
    if not (template.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        return JSONResponse({'detail': '仅支持 .xlsx / .xlsm 表格：%s' % template.filename},
                            status_code=400)
    data = await template.read()
    tmp = os.path.join(tempfile.gettempdir(), 'ws_%s.xlsx' % uuid.uuid4().hex[:8])
    with open(tmp, 'wb') as f:
        f.write(data)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return JSONResponse({'detail': '无法读取表格：%s' % e}, status_code=400)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass
    return {'sheets': sheets}


@app.post('/api/tasks')
async def create_task(files: list[UploadFile] = File(...),
                      mode: str = Form(...),
                      inv_type: str = Form('1'),
                      layout: str = Form('v'),
                      start_cell: str = Form('A1'),
                      template: UploadFile = File(None),
                      sheet_name: str = Form('')):
    if mode not in ('1', '2'):
        return JSONResponse({'detail': 'mode 无效，应为 1 或 2'}, status_code=400)
    if mode == '2' and inv_type not in ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11'):
        return JSONResponse({'detail': 'inv_type 无效，应为 1-11'}, status_code=400)
    if layout not in ('v', 'h'):
        return JSONResponse({'detail': 'layout 无效，应为 v 或 h'}, status_code=400)
    if not files:
        return JSONResponse({'detail': '未上传任何文件'}, status_code=400)
    if template and not (template.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        return JSONResponse({'detail': '表格模板仅支持 .xlsx / .xlsm'}, status_code=400)

    pdfs = []
    for f in files:
        if not (f.filename or '').lower().endswith('.pdf'):
            return JSONResponse({'detail': '仅支持 .pdf 文件：%s' % f.filename},
                                status_code=400)
        pdfs.append((f.filename or 'file.pdf', await f.read()))

    tpl = None
    if template:
        tpl = (template.filename or 'template.xlsx', await template.read())

    task_id = tasks.create_task(pdfs, mode, inv_type, layout, start_cell,
                                template=tpl, sheet_name=sheet_name.strip())
    return {'task_id': task_id, 'files': len(pdfs)}


@app.get('/api/tasks/{task_id}')
def task_status(task_id: str):
    t = tasks.get_task(task_id)
    if not t:
        return JSONResponse({'detail': '任务不存在'}, status_code=404)
    return {
        'status': t['status'],
        'current': t['current'],
        'total': t['total'],
        'message': t['message'],
        'filename': t['filename'],
        'error': t['error'],
    }


@app.get('/api/tasks/{task_id}/download')
def download(task_id: str):
    path = tasks.download_path(task_id)
    if not path:
        return JSONResponse({'detail': '结果不存在或任务未完成'}, status_code=404)
    return FileResponse(path, filename=os.path.basename(path),
                        media_type=_XLSX_MEDIA)


# 托管前端构建产物（dev 阶段 dist 可能不存在，仅由 npm run dev 提供页面）
if os.path.isdir(_FRONT_DIST):
    @app.get('/')
    def index():
        return FileResponse(os.path.join(_FRONT_DIST, 'index.html'))

    assets = os.path.join(_FRONT_DIST, 'assets')
    if os.path.isdir(assets):
        app.mount('/assets', StaticFiles(directory=assets), name='assets')


# 启动单 worker（幂等）
tasks.start_worker()
