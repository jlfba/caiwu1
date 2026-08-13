# -*- coding: utf-8 -*-
"""
PDF 工具：现有发票图片识别 / 发票明细转表格

功能（顶层选择）：
1. 功能：PDF 转图片 + 发票识别
   - 支持一次拖入多个 PDF（用 " " 分隔），逐页渲染为 PNG。
   - 用 OCR（RapidOCR，PP-OCRv4 模型）识别每页发票的四个字段：
     发票号码（右上角）、购买方名称、销售方名称、金额（小写）。
   - 图片按 发票号码_购买方_销售方_金额.png 重新命名。
   - 再选择：1 直接保存到文件夹；2 拖入 Excel 表格，选工作表、起始单元格、方向后批量插入。
   - 插表：图片固定 10cm x 15cm，横向/纵向两种字段排版。
2. 发票明细识别并转 Excel（v2.0.0 新增）
   - 从英文发票 PDF 中识别 INVOICE 号码（右侧）、TRACKING NO.（下一行）
     和 DATE / DESCRIPTION / TAX / QTY / RATE / AMOUNT 明细行。
   - 有文字层的 PDF 用 PyMuPDF 原生坐标，扫描件自动回退 OCR（RapidOCR）。
   - 自动合并 DESCRIPTION/TAX/DATE 的换行内容，过滤 TOTAL/SUBTOTAL 等汇总行。
   - 输出 Excel 固定列：发票号、TRACKING NO.、DATE、DESCRIPTION、TAX、QTY、RATE、AMOUNT；
     发票号和追踪编号按明细行重复；默认保存到首个 PDF 目录 发票明细表.xlsx。
   - 发票类型：1 canexs；2 精准（Accuracy Customs Brokers）；3 创时亚马逊卡派；4 创时卡派；5 创时清关费。

使用：
    python pdf转图片.py
"""

import os
import sys
import re
import datetime

import fitz  # PyMuPDF

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# 渲染分辨率（DPI）：150 → 120，OCR 检测模型耗时随图片面积下降（约快 1.2 倍）
# 发票四字段（号码/名称/金额）字号较大，120 DPI 仍可清晰识别；若个别发票小字识别失败可调回 150
RENDER_DPI = 120

CELL_RE = re.compile(r'^([A-Za-z]+)(\d+)$')

# 字段名称顺序
FIELD_LABELS = ('发票号码', '购买方', '销售方', '金额')


# ---------------------------------------------------------------------------
# 控制台颜色
# ---------------------------------------------------------------------------
# ANSI 转义序列：交互提示用加粗亮绿色，便于与输出信息区分
COLOR_PROMPT = '\033[1;92m'  # 加粗亮绿色（提示文字）
COLOR_RESET = '\033[0m'      # 复位
# 仅当标准输出是终端时启用颜色，避免重定向/管道出现乱码
_USE_COLOR = sys.stdout.isatty()


def init_console_color():
    """Windows 控制台启用 ANSI 转义序列（VT 处理），其他平台无需处理。"""
    if _USE_COLOR and os.name == 'nt':
        try:
            import ctypes
            kernel32 = ctypes.windll.kernel32
            h = kernel32.GetStdHandle(-11)
            mode = ctypes.c_ulong()
            if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
                kernel32.SetConsoleMode(h, mode.value | 0x0004)  # ENABLE_VIRTUAL_TERMINAL_PROCESSING
        except Exception:
            pass


# ---------------------------------------------------------------------------
# 输入工具
# ---------------------------------------------------------------------------
# 匹配拖入路径：优先引号包裹（路径可含空格），否则连续非空白
PATH_RE = re.compile(r'"([^"]*)"|(\S+)')

# 预读缓冲：多行粘贴时，非路径行先暂存，让下一个问题优先读取
_PUTBACK = []


def _ask(prompt):
    """读一行；若预读缓冲有内容则优先取缓冲，否则阻塞等待输入。"""
    if _PUTBACK:
        return _PUTBACK.pop(0)
    if _USE_COLOR:
        prompt = COLOR_PROMPT + prompt + COLOR_RESET
    return input(prompt)


def is_path_like(s):
    """粗略判断一行是否为文件路径（存在该文件，或含路径分隔符/盘符）。"""
    if not s:
        return False
    if os.path.isfile(s):
        return True
    if '/' in s or '\\' in s or ':' in s:
        return True
    return False


def read_paths(prompt):
    """读取一行输入，按 Windows 拖拽形态拆分成多个路径。
    兼容多种形态：
    - 带引号单个："d:/a.pdf"
    - 带引号多个（路径含空格）："d:/a.pdf" "d:/b c.pdf"
    - 无引号多个（路径无空格，空格分隔）：d:/a.pdf d:/b.pdf
    - 输入 'c' 时改为从剪贴板读取全部路径（多选复制，绕过经典终端多选拖入只插第一个的限制）
    - Ctrl+V 多行粘贴：每行一个路径，自动连续读取；非路径行（如模式答案 1/2）暂存给下一个问题
    """
    first = _ask(prompt).strip()
    if first.lower() in ('c', 'cb', 'clip', '粘贴'):
        return read_clipboard_paths()
    paths = []
    if is_path_like(first):
        paths.append(first)
    else:
        return []  # 空行或非路径，交由上层重试
    # 继续读取后续路径行（支持 Ctrl+V 多行粘贴）
    while True:
        nxt = _ask('').strip()
        if not nxt:
            break
        if is_path_like(nxt):
            paths.append(nxt)
        else:
            _PUTBACK.append(nxt)
            break
    return paths


def _read_clipboard_hdrop():
    """读取 Windows 剪贴板的文件列表（CF_HDROP，资源管理器复制文件时写入）。
    返回完整路径列表；剪贴板没有文件列表时返回 []。
    多选复制时文本格式只含文件名，完整路径在 CF_HDROP 里，所以优先读它。
    """
    if os.name != 'nt':
        return []
    try:
        import ctypes
        CF_HDROP = 15
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        # 关键：64 位 Windows 上 HGLOBAL/HANDLE 是指针宽度，
        # 必须显式声明 restype/argtypes，否则 ctypes 按 32 位截断句柄，
        # 后续 GlobalSize/GlobalLock 全失败，文件列表读不到。
        user32.OpenClipboard.argtypes = [ctypes.c_void_p]
        user32.GetClipboardData.restype = ctypes.c_void_p
        kernel32.GlobalSize.restype = ctypes.c_size_t
        kernel32.GlobalSize.argtypes = [ctypes.c_void_p]
        kernel32.GlobalLock.restype = ctypes.c_void_p
        kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
        kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]
        if not user32.OpenClipboard(None):
            return []
        try:
            handle = user32.GetClipboardData(CF_HDROP)
            if not handle:
                return []
            size = kernel32.GlobalSize(handle)
            if not size:
                return []
            p = kernel32.GlobalLock(handle)
            if not p:
                return []
            try:
                data = ctypes.string_at(p, size)
            finally:
                kernel32.GlobalUnlock(handle)
            # DROPFILES 结构：pFiles(4) pt(8) fNC(4) fWide(4)
            offset = int.from_bytes(data[0:4], 'little')
            fWide = bool(int.from_bytes(data[16:20], 'little'))
            raw = data[offset:]
            text = raw.decode('utf-16-le', errors='ignore') if fWide \
                else raw.decode('cp936', errors='ignore')
            return [s.strip() for s in text.split('\x00') if s.strip()]
        finally:
            user32.CloseClipboard()
    except Exception:
        return []


def read_clipboard_paths():
    """从剪贴板读取文件路径列表。
    优先读资源管理器复制文件产生的文件列表（CF_HDROP，带完整路径）；
    否则从文本里解析，但只保留"像真实路径"的条目（文件存在，或带盘符/路径分隔符），
    避免把终端里的普通文字（如下一条要复制的 PDF 名）误当路径。
    """
    paths = _read_clipboard_hdrop()
    if paths:
        return paths
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        root.update()
        try:
            data = root.clipboard_get()
        except tk.TclError:
            return []
        finally:
            root.destroy()

        paths = []
        for line in data.splitlines():
            line = line.strip()
            if not line:
                continue
            for m in PATH_RE.finditer(line):
                p = m.group(1) if m.group(1) is not None else m.group(2)
                p = p.strip()
                if not p:
                    continue
                # 只保留像真实路径的：文件存在，或带盘符/路径分隔符
                if os.path.isfile(p) or ':' in p or '/' in p or '\\' in p:
                    paths.append(p)
        return paths
    except Exception:
        return []


def col_letter_to_num(letters):
    """列字母转列号（A=1, B=2, ... AA=27）。"""
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def col_num_to_letter(n):
    """列号转列字母（1=A）。"""
    s = ''
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(ord('A') + r) + s
    return s


def parse_cell(cell):
    """把单元格坐标（如 C5）解析为 (列号, 行号)，非法返回 None。"""
    m = CELL_RE.match(cell.strip())
    if not m:
        return None
    return col_letter_to_num(m.group(1)), int(m.group(2))


def sanitize(name):
    """去掉文件名中的非法字符。"""
    return re.sub(r'[\\/:*?"<>|\r\n]', '_', name).strip()


# ---------------------------------------------------------------------------
# PDF 转图片
# ---------------------------------------------------------------------------
def pdf_to_images(pdf_path, out_dir, start_index=0, dpi=RENDER_DPI, progress_cb=None):
    """把 PDF 每一页渲染为 PNG，返回 (图片文件列表, 全局页计数)。
    progress_cb(stage, done, total)：每渲染完一页回调一次，供网页版显示进度（可选）。"""
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError('找不到文件：%s' % pdf_path)

    base, ext = os.path.splitext(pdf_path)
    if ext.lower() != '.pdf':
        raise ValueError('不是 PDF 文件：%s' % pdf_path)

    os.makedirs(out_dir, exist_ok=True)

    doc = fitz.open(pdf_path)
    images = []
    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)

    try:
        total = doc.page_count
        print('  %s：共 %d 页' % (os.path.basename(pdf_path), total))
        for i, page in enumerate(doc):
            pix = page.get_pixmap(matrix=mat, alpha=False)
            # 全局序号加两位，防止多 PDF 同名覆盖
            seq = start_index + i + 1
            img_name = '%04d-%s-第%d页.png' % (seq, os.path.basename(base), i + 1)
            img_path = os.path.join(out_dir, img_name)
            pix.save(img_path)
            images.append(img_path)
            print('    [%d/%d] %s' % (i + 1, total, img_name))
            if progress_cb is not None:
                progress_cb('render', i + 1, total)
    finally:
        doc.close()

    return images, start_index + total


# ---------------------------------------------------------------------------
# OCR 发票识别
# ---------------------------------------------------------------------------
_ocr = None


def get_ocr():
    """惰性初始化 OCR（全局单例，避免重复加载模型）。

    用 RapidOCR（onnxruntime）替代 PaddleOCR：
    - PaddlePaddle 官方 wheel 要求 AVX 指令集，在无 AVX 的 CPU（如 Pentium）上必崩（SIGILL）
    - RapidOCR 跑同一套 PP-OCRv4 ONNX 模型（wheel 内置，无需下载），onnxruntime 支持 SSE4.2 老 CPU
    """
    global _ocr
    if _ocr is None:
        from rapidocr_onnxruntime import RapidOCR
        _ocr = RapidOCR()
    return _ocr


def ocr_lines(image_path):
    """对图片做 OCR，返回带坐标的文本条目列表。
    每条为 dict：{text, cx, cy, w, h, aspect}。
    RapidOCR 返回 result = [[box(4x2), text, score], ...]，box 是四点坐标。
    """
    result, _ = get_ocr()(image_path)
    items = []
    if not result:
        return items
    for box, text, _score in result:
        text = text.strip()
        if not text:
            continue
        xs = [float(p[0]) for p in box]
        ys = [float(p[1]) for p in box]
        x1, x2 = min(xs), max(xs)
        y1, y2 = min(ys), max(ys)
        w, h = x2 - x1, y2 - y1
        items.append({'text': text,
                      'cx': (x1 + x2) / 2,
                      'cy': (y1 + y2) / 2,
                      'w': w,
                      'h': h,
                      'aspect': w / h if h > 0 else 99.0})
    return items


def extract_invoice_fields(image_path):
    """识别发票四字段，返回 dict：no/buyer/seller/amount。失败字段用 '未知'。
    适配两种布局：
    - 上下布局（标题在上、名称在下，同一 x 列）
    - 左右分栏（"购买方信息/销售方信息"为竖排标题，名称在标题附近按 x 区分）
    """
    fields = {'no': '未知', 'buyer': '未知', 'seller': '未知', 'amount': '未知'}
    try:
        items = ocr_lines(image_path)
    except Exception as e:
        print('    OCR 失败：%s' % e)
        return fields

    # 1. 发票号码（右上角，含"发票号码"）
    for it in items:
        if '发票号码' in it['text']:
            m = re.search(r'[：:]\s*([0-9A-Za-z\-]+)', it['text'])
            if m:
                fields['no'] = m.group(1)
            break

    # 2. 购买方 / 销售方 名称
    # 找所有"名称："条目
    name_items = [it for it in items if re.search(r'名称\s*[：:]', it['text'])]

    def find_name(header_kw):
        # 找标题条目（含关键词，优先竖排 aspect<1）
        headers = [it for it in items if header_kw in it['text']]
        if not headers:
            return '未知'
        header = min(headers, key=lambda it: it['aspect'])  # 最竖的那条
        if not name_items:
            return '未知'
        # 取与标题水平距离最近且垂直距离合理的"名称："条目
        def score(it):
            dx = abs(it['cx'] - header['cx'])
            dy = abs(it['cy'] - header['cy'])
            return dx + dy
        best = min(name_items, key=score)
        m = re.search(r'名称\s*[：:]\s*(.+)', best['text'])
        if m and m.group(1).strip():
            return m.group(1).strip()
        return '未知'

    fields['buyer'] = find_name('购买方')
    fields['seller'] = find_name('销售方')

    # 3. 金额（小写）
    for it in items:
        if '小写' in it['text']:
            # 兼容 OCR 可能输出的中文全角逗号（，）与英文半角逗号（,）
            m = re.search(r'(?:¥\s*)?([\d,，]+\.\d{1,2})', it['text'])
            if m:
                fields['amount'] = m.group(1).replace(',', '').replace('，', '')
            break

    return fields


# ---------------------------------------------------------------------------
# 图片重命名（发票号_购买方_销售方_金额）
# ---------------------------------------------------------------------------
def rename_with_fields(img_path, fields):
    """按四字段重命名图片，返回新路径。若重名则追加序号。"""
    name = '%s_%s_%s_%s.png' % (fields['no'], fields['buyer'],
                                fields['seller'], fields['amount'])
    new_path = os.path.join(os.path.dirname(img_path), sanitize(name))
    if os.path.abspath(new_path) == os.path.abspath(img_path):
        return new_path
    # 若目标已存在，追加序号
    if os.path.exists(new_path):
        base, ext = os.path.splitext(new_path)
        i = 2
        while os.path.exists('%s(%d)%s' % (base, i, ext)):
            i += 1
        new_path = '%s(%d)%s' % (base, i, ext)
    os.rename(img_path, new_path)
    return new_path


# ---------------------------------------------------------------------------
# PDF 发票明细识别
# ---------------------------------------------------------------------------
DETAIL_HEADERS = ('DATE', 'DESCRIPTION', 'TAX', 'QTY', 'RATE', 'AMOUNT')
DETAIL_OUTPUT_HEADERS = ('发票号', 'TRACKING NO.', 'DATE', 'DESCRIPTION',
                        'TAX', 'QTY', 'RATE', 'AMOUNT')

# 精准发票（Accuracy Customs Brokers，美国清关行）输出列：
# Invoice Number / Master B/L No / Containers 三个字段值按 Description 行重复
JINGZHUN_OUTPUT_HEADERS = ('发票号', 'Master B/L No', 'Containers',
                           'Description', 'Amount')

# 创时亚马逊卡派发票输出列：Invoice Number / Reference 按明细行重复，
# Description 列内容拆成两列——带 // 的行放 Description，其余行放 Description(1)
CHUANGSHI_OUTPUT_HEADERS = ('Invoice Number', 'Reference', 'Description',
                            'Description(1)', 'Quantity', 'Unit Price', 'Amount GBP')

# 创时卡派发票输出列：不带 Description(1)，只有一个 Description 列
CHUANGSHI_CAR_OUTPUT_HEADERS = ('Invoice Number', 'Reference', 'Description',
                                'Quantity', 'Unit Price', 'Amount GBP')

# 创时清关费发票输出列：同创时卡派，6 列无 Description(1)
CHUANGSHI_CLEARANCE_OUTPUT_HEADERS = ('Invoice Number', 'Reference', 'Description',
                                      'Quantity', 'Unit Price', 'Amount GBP')


def _compact_text(text):
    """统一 OCR/原生文字中的空格和标点，便于匹配英文标签。"""
    return re.sub(r'[^A-Z0-9]', '', text.upper())


def pdf_native_items(pdf_path, page):
    """读取 PDF 原生文字并转换为与 OCR 相同的坐标条目。"""
    items = []
    for word in page.get_text('words'):
        x1, y1, x2, y2, text = word[:5]
        text = str(text).strip()
        if text:
            items.append({'text': text, 'cx': (x1 + x2) / 2,
                          'cy': (y1 + y2) / 2, 'w': x2 - x1,
                          'h': y2 - y1,
                          'aspect': (x2 - x1) / (y2 - y1) if y2 > y1 else 99})
    return items


def _detail_page_items(pdf_path, page, page_no):
    """优先使用 PDF 文字层；没有文字层时渲染本页并用 OCR。"""
    items = pdf_native_items(pdf_path, page)
    if items:
        return items
    zoom = RENDER_DPI / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    tmp = os.path.join(os.path.dirname(pdf_path),
                       '.pdf_detail_%s_%d.png' % (os.getpid(), page_no))
    try:
        pix.save(tmp)
        return ocr_lines(tmp)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def _group_detail_lines(items, y_tolerance=None):
    """按文字块的纵坐标合并成视觉行，保留从左到右顺序。"""
    if not items:
        return []
    items = sorted(items, key=lambda x: (x['cy'], x['cx']))
    if y_tolerance is None:
        heights = [x['h'] for x in items if x.get('h', 0) > 0]
        y_tolerance = max(3.0, (sum(heights) / len(heights) if heights else 10) * 0.7)
    lines = []
    for item in items:
        target = None
        for line in reversed(lines[-2:]):
            if abs(item['cy'] - line['cy']) <= y_tolerance:
                target = line
                break
        if target is None:
            target = {'cy': item['cy'], 'items': []}
            lines.append(target)
        target['items'].append(item)
        target['cy'] = sum(x['cy'] for x in target['items']) / len(target['items'])
    for line in lines:
        line['items'].sort(key=lambda x: x['cx'])
        line['text'] = ' '.join(x['text'] for x in line['items'])
    return lines


# 明细行的日期匹配：DATE 列只要含日期就算新一行（TAX/DESCRIPTION 换行即使含数字也不会误判）
DETAIL_DATE_RE = re.compile(
    r'((?:19|20)\d{2}[.\-/]\d{1,2}[.\-/]\d{1,2}|'
    r'\d{1,2}[.\-/]\d{1,2}[.\-/](?:19|20)\d{2}|'
    r'\d{1,2}[.\-/]\d{2,4}|'
    r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?[ \t]+\d{1,2})')


def _find_line(lines, compact_key):
    """按合并后的英文标签找行。标签可能是多个词（如 TRACKING NO.）。"""
    for ln in lines:
        if compact_key in _compact_text(ln['text']):
            return ln
    return None


def _line_right_of(line, anchor, drop=()):
    """anchor 右侧同一行的内容，去掉 drop 中的纯标签词后拼接。"""
    parts = [x['text'] for x in line['items'] if x['cx'] > anchor['cx']]
    parts = [p for p in parts if _compact_text(p) not in drop]
    return ' '.join(parts).strip()


def _extract_invoice_no(lines):
    """发票号码：INVOICE 右侧（同行）的内容，去掉 NO./Nº 等标签词。"""
    ln = _find_line(lines, 'INVOICE')
    if ln is None:
        return '未知'
    anchor = next((x for x in ln['items']
                   if _compact_text(x['text']).startswith('INVOICE')), None)
    if anchor is not None:
        raw = _line_right_of(ln, anchor, drop=('NO', 'NO.', 'N', 'NO:'))
        if raw:
            return raw
        m = re.search(r'([A-Za-z0-9][A-Za-z0-9\-]{2,})', anchor['text'])
        if m:
            return m.group(1)
    return '未知'


def _extract_tracking_no(lines):
    """追踪编号：TRACKING NO. 下方最近一行的内容（同行右侧优先）。"""
    ln = _find_line(lines, 'TRACKINGNO')
    if ln is None:
        return '未知'
    anchor = next((x for x in ln['items']
                   if _compact_text(x['text']).startswith('TRACKING')), None)
    if anchor is not None:
        raw = _line_right_of(ln, anchor, drop=('NO', 'NO.', 'N'))
        if raw:
            return raw
    below = [x for x in lines if x['cy'] > ln['cy'] + 1]
    if not below:
        return '未知'
    nxt = min(below, key=lambda x: x['cy'])
    return nxt['text'].strip() or '未知'


def _find_header_band(items):
    """找明细表头：同一横带里包含最多不同列标签的一组表头词。
    返回 {label: item}；找不到（不足 4 列）返回 None。
    用完全匹配（去空格标点后），避免把页面上部的发票日期 DATE 字段等误当表头。
    """
    keys = {label: _compact_text(label) for label in DETAIL_HEADERS}
    candidates = []
    for it in items:
        t = _compact_text(it['text'])
        for label, key in keys.items():
            if t == key:
                candidates.append((label, it))
                break
    if not candidates:
        return None
    heights = [x['h'] for x in items if x.get('h', 0) > 0]
    tol = max(3.0, (sum(heights) / len(heights) if heights else 10) * 0.7)
    bands = []
    for label, it in candidates:
        placed = None
        for b in bands:
            if abs(it['cy'] - b['cy']) <= tol:
                placed = b
                break
        if placed is None:
            placed = {'cy': it['cy'], 'items': []}
            bands.append(placed)
        placed['items'].append((label, it))
        placed['cy'] = sum(x[1]['cy'] for x in placed['items']) / len(placed['items'])
    best = max(bands, key=lambda b: len({lab for lab, _ in b['items']}))
    found = {}
    for label, it in best['items']:
        if label not in found or it['cx'] < found[label]['cx']:
            found[label] = it
    if len(found) < 4:
        return None
    return found


def extract_detail_rows(items):
    """从一页坐标文字中提取发票号、追踪号和明细行。"""
    lines = _group_detail_lines(items)
    invoice = _extract_invoice_no(lines)
    tracking = _extract_tracking_no(lines)
    found = _find_header_band(items)
    if found is None:
        return invoice, tracking, []
    # 表头按左边缘排序列边界，词条按中心 cx 归列。
    # DATE 列右边界取 DESCRIPTION 表头左边缘，日期文字横向越界也不会被误并入 DESCRIPTION。
    ordered = sorted(found.items(), key=lambda kv: kv[1]['cx'] - kv[1]['w'] / 2)  # [(label, item)]
    lefts = [it['cx'] - it['w'] / 2 for _, it in ordered]
    label_col = {label: idx for idx, label in enumerate(DETAIL_HEADERS)}
    bounds = [float('-inf')] + lefts[1:] + [float('inf')]
    col_index = [label_col[lab] for lab, _ in ordered]
    # 数据区：位于表头底部下方（表头底部=cy+字高/2，避免第一行紧贴表头被误排除）。
    header_bottom = max(it['cy'] + it['h'] / 2 for _, it in ordered)
    data_lines = [ln for ln in lines if ln['cy'] > header_bottom + 2]
    rows = []
    stop_words = ('SUBTOTAL', 'TOTAL', 'BALANCE', 'PAYMENT', 'THANK')
    for line in data_lines:
        if any(word in _compact_text(line['text']) for word in stop_words):
            break
        cells = [''] * len(DETAIL_HEADERS)
        for item in line['items']:
            col = next((ci for ci in range(len(bounds) - 1)
                        if bounds[ci] <= item['cx'] < bounds[ci + 1]), None)
            if col is not None:
                cells[col_index[col]] = (cells[col_index[col]] + ' ' + item['text']).strip()
        # DATE 列含日期代表新明细；否则是 DESCRIPTION/TAX/DATE 的换行，合并到上一行。
        is_new = bool(DETAIL_DATE_RE.search(cells[0])) or (not rows and any(cells))
        if is_new:
            rows.append(cells)
        elif rows and any(cells):
            for i, value in enumerate(cells):
                if value:
                    rows[-1][i] = (rows[-1][i] + ' ' + value).strip()
    result = []
    for cells in rows:
        if not any(cells):
            continue
        result.append([invoice, tracking] + cells)
    return invoice, tracking, result



def extract_detail_from_pdfs(pdf_paths):
    """批量识别 PDF，返回明细行和处理统计。续页自动继承上一页的发票号/追踪编号。"""
    all_rows, pages, skipped = [], 0, 0
    for pdf_path in pdf_paths:
        if not os.path.isfile(pdf_path):
            print('  找不到文件，跳过：%s' % pdf_path)
            skipped += 1
            continue
        doc = fitz.open(pdf_path)
        last_invoice = last_tracking = '未知'
        try:
            for page_no, page in enumerate(doc, 1):
                pages += 1
                items = _detail_page_items(pdf_path, page, page_no)
                invoice, tracking, rows = extract_detail_rows(items)
                if invoice != '未知':
                    last_invoice = invoice
                if tracking != '未知':
                    last_tracking = tracking
                for row in rows:
                    if row[0] == '未知':
                        row[0] = last_invoice
                    if row[1] == '未知':
                        row[1] = last_tracking
                    all_rows.append(row)
        finally:
            doc.close()
    return all_rows, pages, skipped


def _next_output_path(directory, filename):
    path = os.path.join(directory, filename)
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    index = 2
    while os.path.exists('%s(%d)%s' % (base, index, ext)):
        index += 1
    return '%s(%d)%s' % (base, index, ext)


def to_number(s):
    """把可解析的数字文本转成数字（int/float），去货币符号和千分位；转不了原样返回。"""
    if not isinstance(s, str):
        return s
    t = s.strip().replace(',', '').replace('$', '').replace('¥', '').replace('￥', '').replace('%', '')
    if not t:
        return s
    try:
        if re.fullmatch(r'[+-]?\d+', t):
            return int(t)
        if re.fullmatch(r'[+-]?\d+\.\d+', t):
            return float(t)
    except (ValueError, OverflowError):
        pass
    return s


def write_detail_excel(rows, output_path, headers=None, numeric_cols=None, widths=None, text_cols=None, zero_pad_cols=None):
    """将发票明细写入 Excel，并开启筛选、冻结和换行。
    headers：输出表头；numeric_cols：转成纯数字的列序号（可转的才转）；
    widths：各列宽；text_cols：强制按文本写入并设文本格式的列序号；
    zero_pad_cols：转成数字 + 自定义格式补前导 0 的列序号（如发票号 0098726 →
    存数字 98726、格式 0000000，显示仍是 0098726，且不会出现"数字以文本存储"的绿三角）。
    缺省按 canexs 明细表（发票号/QTY/RATE/AMOUNT 转数字）。
    """
    if not OPENPYXL_OK:
        raise RuntimeError('未安装 openpyxl，无法写入 Excel。请运行：pip install openpyxl')
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    if headers is None:
        headers = DETAIL_OUTPUT_HEADERS
    if numeric_cols is None:
        numeric_cols = {0, 5, 6, 7}
    if widths is None:
        widths = [18, 24, 28, 45, 25, 12, 14, 16]
    if text_cols is None:
        text_cols = set()
    if zero_pad_cols is None:
        zero_pad_cols = set()
    wb = load_workbook(output_path) if os.path.exists(output_path) else None
    if wb is None:
        from openpyxl import Workbook
        wb = Workbook()
    ws = wb.active
    ws.title = '发票明细'
    ws.delete_rows(1, ws.max_row)
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='113584', end_color='113584', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # 零填充列的最长位数（按原始文本长度，保证显示既不缺也不多补 0）
    pad_lens = {}
    for i in zero_pad_cols:
        lens = [len(str(row[i])) for row in rows if row[i] not in (None, '')]
        pad_lens[i] = max(lens) if lens else 0
    numeric_set = numeric_cols | zero_pad_cols
    for row in rows:
        ws.append([to_number(v) if i in numeric_set else v for i, v in enumerate(row)])
    # 文本列：显式设文本格式，值若被转成数字也还原为字符串（保住前导 0）
    for i in text_cols:
        for cell in ws[get_column_letter(i + 1)][1:]:
            if isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)
            cell.number_format = '@'
    # 零填充数字列：真实值是数字，自定义格式补前导 0（无绿三角，显示仍带 0）
    for i in zero_pad_cols:
        n = pad_lens.get(i, 0)
        if n <= 0:
            continue
        for cell in ws[get_column_letter(i + 1)][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0' * n
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path


def canexs_mode(pdf_paths):
    """canexs 发票：提取 INVOICE / TRACKING NO. / 明细行，输出到 canexs发票-日期 文件夹。"""
    print('识别 canexs 发票明细（文字层优先，扫描件自动使用 OCR）…')
    rows, pages, skipped = extract_detail_from_pdfs(pdf_paths)
    if skipped:
        print('有 %d 个文件找不到，已跳过。' % skipped)
        print('提示：请在资源管理器选中文件按 Ctrl+C 复制，再输入 c（这样才带完整路径）；或直接把文件拖入窗口。')
    if not rows:
        print('没有识别到任何明细，未生成 Excel。')
        return
    folder = os.path.join(os.path.dirname(os.path.abspath(pdf_paths[0])),
                          'canexs发票-%s' % datetime.datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(folder, exist_ok=True)
    output = _next_output_path(folder, '发票明细表.xlsx')
    write_detail_excel(rows, output)
    print('识别完成：共处理 %d 页，提取 %d 行明细。' % (pages, len(rows)))
    print('Excel 已保存：%s' % output)


# 精准发票字段值：标签在上一行、值在下一行同列。取标签中心 ± _JZ_COL_BAND 内的词作为该列值。
_JZ_COL_BAND = 65.0


def _jz_label(lines, line_key, word_key):
    """在标签行里找锚点词：行的紧凑串含 line_key，词紧凑串等于 word_key。
    同一行可能有多个匹配词（如 Invoice Number / Invoice Date 都有 INVOICE），取最左那个。
    返回 (anchor_item, label_line) 或 (None, None)。"""
    for ln in lines:
        if line_key in _compact_text(ln['text']):
            cands = [w for w in ln['items'] if _compact_text(w['text']) == word_key]
            if cands:
                return min(cands, key=lambda w: w['cx']), ln
    return None, None


def _jz_value(lines, header_cy, anchor, label_line):
    """取标签下方最近一行的同列内容（值与标签左对齐，中心距离 ≤ 波段）。"""
    if anchor is None:
        return '未知'
    below = [ln for ln in lines if ln['cy'] > label_line['cy'] + 2 and ln['cy'] < header_cy - 2]
    if not below:
        return '未知'
    nxt = min(below, key=lambda ln: ln['cy'])
    vals = [w for w in nxt['items'] if abs(w['cx'] - anchor['cx']) <= _JZ_COL_BAND]
    if not vals:
        return '未知'
    vals.sort(key=lambda w: w['cx'])
    return ' '.join(w['text'] for w in vals).strip()


def _jz_table(lines, header_line):
    """Description|Amount 两列表：表头下方每行一条，行项在左、金额右对齐，遇 TOTAL/NOTES 停止。"""
    desc_hdr = next((w for w in header_line['items']
                     if _compact_text(w['text']) == 'DESCRIPTION'), None)
    amt_hdr = next((w for w in header_line['items']
                    if _compact_text(w['text']) == 'AMOUNT'), None)
    if desc_hdr is None or amt_hdr is None:
        return []
    threshold = (desc_hdr['cx'] + amt_hdr['cx']) / 2
    rows = []
    below = sorted((ln for ln in lines if ln['cy'] > header_line['cy'] + 2),
                   key=lambda ln: ln['cy'])
    for ln in below:
        comp = _compact_text(ln['text'])
        if 'TOTAL' in comp or 'NOTES' in comp or 'NOTICE' in comp:
            break
        desc = ' '.join(w['text'] for w in ln['items'] if w['cx'] < threshold)
        amt = ' '.join(w['text'] for w in ln['items'] if w['cx'] >= threshold)
        if desc or amt:
            rows.append((desc, amt))
    return rows


def extract_jingzhun_page(items):
    """精准（Accuracy Customs Brokers）发票单页：字段值 + Description|Amount 表行。
    返回 ({invoice_no, master_bl, containers}, [(desc, amount), ...])。"""
    if not items:
        return {}, []
    lines = _group_detail_lines(items)
    # 表头行：同时含 DESCRIPTION 和 AMOUNT（避免把上方的 Commercial Description 标签当表头）
    header_line = None
    for ln in lines:
        comp = _compact_text(ln['text'])
        if 'DESCRIPTION' in comp and 'AMOUNT' in comp:
            header_line = ln
            break
    if header_line is None:
        return {}, []
    header_cy = header_line['cy']
    field_lines = [ln for ln in lines if ln['cy'] < header_cy - 2]
    fields = {}
    for key, line_key, word_key in (('invoice_no', 'INVOICENUMBER', 'INVOICE'),
                                    ('master_bl', 'MASTERBLNO', 'MASTER'),
                                    ('containers', 'CONTAINERS', 'CONTAINERS')):
        anchor, label_line = _jz_label(field_lines, line_key, word_key)
        fields[key] = _jz_value(field_lines, header_cy, anchor, label_line)
    return fields, _jz_table(lines, header_line)


def extract_jingzhun_from_pdfs(pdf_paths):
    """批量识别精准发票，返回明细行和处理统计。续页自动继承上一页的字段值。"""
    all_rows, pages, skipped = [], 0, 0
    for pdf_path in pdf_paths:
        if not os.path.isfile(pdf_path):
            print('  找不到文件，跳过：%s' % pdf_path)
            skipped += 1
            continue
        doc = fitz.open(pdf_path)
        last = {}
        try:
            for page_no, page in enumerate(doc, 1):
                pages += 1
                items = _detail_page_items(pdf_path, page, page_no)
                fields, rows = extract_jingzhun_page(items)
                merged = dict(last)
                merged.update({k: v for k, v in fields.items() if v != '未知'})
                last = merged
                for desc, amt in rows:
                    all_rows.append([merged.get('invoice_no', '未知'),
                                     merged.get('master_bl', '未知'),
                                     merged.get('containers', '未知'),
                                     desc, amt])
        finally:
            doc.close()
    return all_rows, pages, skipped


def jingzhun_mode(pdf_paths):
    """精准发票：提取 Invoice Number / Master B/L No / Containers / Description / Amount，
    输出到 精准发票-日期 文件夹。"""
    print('识别精准发票明细（Accuracy Customs Brokers，文字层优先，扫描件自动使用 OCR）…')
    rows, pages, skipped = extract_jingzhun_from_pdfs(pdf_paths)
    if skipped:
        print('有 %d 个文件找不到，已跳过。' % skipped)
        print('提示：请在资源管理器选中文件按 Ctrl+C 复制，再输入 c（这样才带完整路径）；或直接把文件拖入窗口。')
    if not rows:
        print('没有识别到任何明细，未生成 Excel。')
        return
    folder = os.path.join(os.path.dirname(os.path.abspath(pdf_paths[0])),
                          '精准发票-%s' % datetime.datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(folder, exist_ok=True)
    output = _next_output_path(folder, '发票明细表.xlsx')
    write_detail_excel(rows, output, headers=JINGZHUN_OUTPUT_HEADERS,
                       numeric_cols={4}, zero_pad_cols={0}, widths=[16, 26, 18, 50, 16])
    print('识别完成：共处理 %d 页，提取 %d 行明细。' % (pages, len(rows)))
    print('Excel 已保存：%s' % output)


def _jz_reference(lines, header_cy, anchor, label_line):
    """Reference 标签下方、表格表头之前的所有内容行，换行合并成一个单元格。"""
    if anchor is None:
        return '未知'
    below = [ln for ln in lines if ln['cy'] > label_line['cy'] + 2
             and ln['cy'] < header_cy - 2]
    parts = [' '.join(w['text'] for w in ln['items']).strip() for ln in below]
    parts = [p for p in parts if p]
    return '\n'.join(parts) or '未知'


def _desc_amazon(desc_lines):
    """创时亚马逊卡派：带 // 的行放 Description 列，其余行放 Description(1) 列。"""
    slash = [d for d in desc_lines if '//' in d]
    other = [d for d in desc_lines if '//' not in d]
    return '\n'.join(slash), '\n'.join(other)


# 创时卡派：货物明细行特征（重量/尺寸、PLTS 卡板、占板位置），单独成行
_GOODS_LINE_RE = re.compile(r'^\d+kg|^\d+\s*PLTS|^占\d')


def _is_goods_line(s):
    return bool(_GOODS_LINE_RE.search(s))


def _desc_car(desc_lines):
    """创时卡派：// 链（可能跨视觉行换行）去行断拼接后按 // 切分，每段一行；
    货物明细行（kg/尺寸、PLTS、占板）单独一行；全部合进 Description 单元格。"""
    lines = [d.strip() for d in desc_lines if d.strip()]
    chain, goods = '', []
    for j, line in enumerate(lines):
        if '//' not in line and _is_goods_line(line):
            goods = lines[j:]
            break
        chain += line  # 去行断拼接（G51+3EB→G513EB），保留行内空格
    if '//' in chain:
        parts = [p.strip() for p in chain.split('//') if p.strip()]
        out = parts + goods
    else:
        out = ([chain] if chain else []) + goods
    return '\n'.join(out), ''


def _chuangshi_table(lines, header_line, desc_parser=_desc_amazon):
    """创时亚马逊卡派/创时卡派四列明细表 Description|Quantity|Unit Price|Amount GBP。

    表头下方逐行按列归类；出现 Quantity/Unit Price/Amount 的行 = 新一行明细的起始，
    其后的 Description 行（可能多行）合并进该行，直到下一价格行或汇总行。
    desc_parser(desc_lines) -> (desc, desc1)，决定 Description 两列如何拆分。
    返回每行 [desc, desc1, qty, unit, amt]。
    """
    def hdr(key):
        for w in header_line['items']:
            if _compact_text(w['text']) == key:
                return w
        return None

    desc_hdr, qty_hdr = hdr('DESCRIPTION'), hdr('QUANTITY')
    unit_hdr, amt_hdr = hdr('UNIT'), hdr('AMOUNT')
    if desc_hdr is None or qty_hdr is None or unit_hdr is None or amt_hdr is None:
        return []
    bounds = [desc_hdr['cx'] - desc_hdr['w'] / 2,
              qty_hdr['cx'] - qty_hdr['w'] / 2,
              unit_hdr['cx'] - unit_hdr['w'] / 2,
              amt_hdr['cx'] - amt_hdr['w'] / 2, float('inf')]
    header_bottom = max(w['cy'] + w['h'] / 2 for w in header_line['items'])
    rows, cur = [], None
    stop_words = ('SUBTOTAL', 'TOTAL', 'BALANCE', 'PAYMENT', 'THANK', 'DUE')
    for ln in lines:
        if ln['cy'] <= header_bottom + 2:
            continue
        comp = _compact_text(ln['text'])
        if any(w in comp for w in stop_words):
            break
        cells = [''] * 4
        for item in ln['items']:
            ci = next((i for i in range(4)
                       if bounds[i] <= item['cx'] < bounds[i + 1]), None)
            if ci is not None:
                cells[ci] = (cells[ci] + ' ' + item['text']).strip()
        desc, qty, unit, amt = cells
        if qty or unit or amt:
            cur = {'desc': [desc] if desc else [], 'qty': qty,
                   'unit': unit, 'amt': amt}
            rows.append(cur)
        elif cur is not None and desc:
            cur['desc'].append(desc)
    result = []
    for r in rows:
        desc, desc1 = desc_parser(r['desc'])
        result.append([desc, desc1, r['qty'], r['unit'], r['amt']])
    return result


def extract_chuangshi_page(items, desc_parser=_desc_amazon):
    """创时亚马逊卡派/创时卡派发票单页：Invoice Number / Reference 字段 + 四列明细行。
    返回 ({invoice_no, reference}, [[desc, desc1, qty, unit, amt], ...])。"""
    if not items:
        return {}, []
    lines = _group_detail_lines(items)
    header_line = None
    for ln in lines:
        comp = _compact_text(ln['text'])
        if ('DESCRIPTION' in comp and 'QUANTITY' in comp
                and 'PRICE' in comp and 'AMOUNT' in comp):
            header_line = ln
            break
    if header_line is None:
        return {}, []
    header_cy = header_line['cy']
    field_lines = [ln for ln in lines if ln['cy'] < header_cy - 2]
    fields = {}
    anchor, label_line = _jz_label(field_lines, 'INVOICENUMBER', 'INVOICE')
    fields['invoice_no'] = _jz_value(field_lines, header_cy, anchor, label_line)
    anchor, label_line = _jz_label(field_lines, 'REFERENCE', 'REFERENCE')
    fields['reference'] = _jz_reference(field_lines, header_cy, anchor, label_line)
    return fields, _chuangshi_table(lines, header_line, desc_parser)


def extract_chuangshi_from_pdfs(pdf_paths):
    """批量识别创时亚马逊卡派发票，返回明细行和处理统计。续页继承上页字段值。"""
    all_rows, pages, skipped = [], 0, 0
    for pdf_path in pdf_paths:
        if not os.path.isfile(pdf_path):
            print('  找不到文件，跳过：%s' % pdf_path)
            skipped += 1
            continue
        doc = fitz.open(pdf_path)
        last = {}
        try:
            for page_no, page in enumerate(doc, 1):
                pages += 1
                items = _detail_page_items(pdf_path, page, page_no)
                fields, rows = extract_chuangshi_page(items)
                merged = dict(last)
                merged.update({k: v for k, v in fields.items() if v != '未知'})
                last = merged
                for r in rows:
                    all_rows.append([merged.get('invoice_no', '未知'),
                                     merged.get('reference', '未知')] + r)
        finally:
            doc.close()
    return all_rows, pages, skipped


def chuangshi_mode(pdf_paths):
    """创时亚马逊卡派发票：Invoice Number / Reference / Description 明细，
    输出到 创时亚马逊卡派发票-日期 文件夹。"""
    print('识别创时亚马逊卡派发票明细（文字层优先，扫描件自动使用 OCR）…')
    rows, pages, skipped = extract_chuangshi_from_pdfs(pdf_paths)
    if skipped:
        print('有 %d 个文件找不到，已跳过。' % skipped)
        print('提示：请在资源管理器选中文件按 Ctrl+C 复制，再输入 c（这样才带完整路径）；或直接把文件拖入窗口。')
    if not rows:
        print('没有识别到任何明细，未生成 Excel。')
        return
    folder = os.path.join(os.path.dirname(os.path.abspath(pdf_paths[0])),
                          '创时亚马逊卡派发票-%s' % datetime.datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(folder, exist_ok=True)
    output = _next_output_path(folder, '发票明细表.xlsx')
    write_detail_excel(rows, output, headers=CHUANGSHI_OUTPUT_HEADERS,
                       numeric_cols={4, 5, 6}, zero_pad_cols=set(),
                       widths=[16, 20, 24, 46, 12, 14, 16])
    print('识别完成：共处理 %d 页，提取 %d 行明细。' % (pages, len(rows)))
    print('Excel 已保存：%s' % output)


def _desc_simple(desc_lines):
    """创时清关费：Description 单列，行内容直接换行合并。"""
    return '\n'.join(desc_lines), ''


def _extract_chuangshi_6col(pdf_paths, desc_parser):
    """创时卡派/创时清关费通用批量识别（输出 6 列，无 Description(1)）。"""
    all_rows, pages, skipped = [], 0, 0
    for pdf_path in pdf_paths:
        if not os.path.isfile(pdf_path):
            print('  找不到文件，跳过：%s' % pdf_path)
            skipped += 1
            continue
        doc = fitz.open(pdf_path)
        last = {}
        try:
            for page_no, page in enumerate(doc, 1):
                pages += 1
                items = _detail_page_items(pdf_path, page, page_no)
                fields, rows = extract_chuangshi_page(items, desc_parser)
                merged = dict(last)
                merged.update({k: v for k, v in fields.items() if v != '未知'})
                last = merged
                for r in rows:
                    all_rows.append([merged.get('invoice_no', '未知'),
                                     merged.get('reference', '未知'),
                                     r[0], r[2], r[3], r[4]])
        finally:
            doc.close()
    return all_rows, pages, skipped


def extract_chuangshi_car_from_pdfs(pdf_paths):
    """批量识别创时卡派发票（Description 为 // 链 + 货物明细行格式）。
    输出行 [invoice, reference, desc, qty, unit, amt]，不带 Description(1) 列。"""
    return _extract_chuangshi_6col(pdf_paths, _desc_car)


def extract_chuangshi_clearance_from_pdfs(pdf_paths):
    """批量识别创时清关费发票（Description 为单行明细）。
    输出行 [invoice, reference, desc, qty, unit, amt]。"""
    return _extract_chuangshi_6col(pdf_paths, _desc_simple)


def chuangshi_car_mode(pdf_paths):
    """创时卡派发票：Invoice Number / Reference / Description 明细（// 链格式），
    输出到 创时卡派发票-日期 文件夹。"""
    print('识别创时卡派发票明细（文字层优先，扫描件自动使用 OCR）…')
    rows, pages, skipped = extract_chuangshi_car_from_pdfs(pdf_paths)
    if skipped:
        print('有 %d 个文件找不到，已跳过。' % skipped)
        print('提示：请在资源管理器选中文件按 Ctrl+C 复制，再输入 c（这样才带完整路径）；或直接把文件拖入窗口。')
    if not rows:
        print('没有识别到任何明细，未生成 Excel。')
        return
    folder = os.path.join(os.path.dirname(os.path.abspath(pdf_paths[0])),
                          '创时卡派发票-%s' % datetime.datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(folder, exist_ok=True)
    output = _next_output_path(folder, '发票明细表.xlsx')
    write_detail_excel(rows, output, headers=CHUANGSHI_CAR_OUTPUT_HEADERS,
                       numeric_cols={3, 4, 5}, zero_pad_cols=set(),
                       widths=[16, 20, 24, 12, 14, 16])
    print('识别完成：共处理 %d 页，提取 %d 行明细。' % (pages, len(rows)))
    print('Excel 已保存：%s' % output)


def chuangshi_clearance_mode(pdf_paths):
    """创时清关费发票：Invoice Number / Reference / Description 明细，
    输出到 创时清关费发票-日期 文件夹。"""
    print('识别创时清关费发票明细（文字层优先，扫描件自动使用 OCR）…')
    rows, pages, skipped = extract_chuangshi_clearance_from_pdfs(pdf_paths)
    if skipped:
        print('有 %d 个文件找不到，已跳过。' % skipped)
        print('提示：请在资源管理器选中文件按 Ctrl+C 复制，再输入 c（这样才带完整路径）；或直接把文件拖入窗口。')
    if not rows:
        print('没有识别到任何明细，未生成 Excel。')
        return
    folder = os.path.join(os.path.dirname(os.path.abspath(pdf_paths[0])),
                          '创时清关费发票-%s' % datetime.datetime.now().strftime('%Y-%m-%d'))
    os.makedirs(folder, exist_ok=True)
    output = _next_output_path(folder, '发票明细表.xlsx')
    write_detail_excel(rows, output, headers=CHUANGSHI_CLEARANCE_OUTPUT_HEADERS,
                       numeric_cols={3, 4, 5}, zero_pad_cols=set(),
                       widths=[16, 20, 34, 12, 14, 16])
    print('识别完成：共处理 %d 页，提取 %d 行明细。' % (pages, len(rows)))
    print('Excel 已保存：%s' % output)


def detail_mode(pdf_paths, inv_type):
    """模式 2：识别发票明细并导出 Excel。
    inv_type: '1' canexs | '2' 精准 | '3' 创时亚马逊卡派 | '4' 创时卡派 | '5' 创时清关费。"""
    if inv_type == '2':
        jingzhun_mode(pdf_paths)
    elif inv_type == '3':
        chuangshi_mode(pdf_paths)
    elif inv_type == '4':
        chuangshi_car_mode(pdf_paths)
    elif inv_type == '5':
        chuangshi_clearance_mode(pdf_paths)
    else:
        canexs_mode(pdf_paths)


# 图片固定 10cm x 15cm。openpyxl 图片锚定单元格左上角、尺寸不受单元格约束，
# 因此把单元格列宽/行高调到正好容纳图片，图片即填满单个单元格。
IMG_W_CM = 15.0
IMG_H_CM = 10.0
IMG_W_PX = int(IMG_W_CM * 96 / 2.54)    # ≈567px
IMG_H_PX = int(IMG_H_CM * 96 / 2.54)    # ≈378px
COL_WIDTH = (IMG_W_PX - 5) / 7.0        # Excel 列宽单位（字符数）
ROW_HEIGHT = IMG_H_PX * 3 / 4.0         # Excel 行高（磅）
TEXT_COL_WIDTH = 40                     # 字段列宽（放得下公司名称）


def images_into_excel(xlsx_path, images, sheet_name=None,
                      start_cell='A1', direction='v'):
    """
    把图片插入指定工作表，每张图占一个单元格（15cm x 10cm，宽:高≈1.5），字段放在图片旁。
    direction：'v' 纵向（图片沿列向下，字段放图片右侧）；'h' 横向（图片沿行向右，字段放图片下方）。
    """
    if not OPENPYXL_OK:
        raise RuntimeError('未安装 openpyxl，无法写入 Excel。请运行：pip install openpyxl')

    wb = load_workbook(xlsx_path)
    if sheet_name is None:
        ws = wb.active
    else:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError('工作表中不存在：%s' % sheet_name)
        ws = wb[sheet_name]

    col, row = parse_cell(start_cell)
    if not col:
        wb.close()
        raise ValueError('无效的单元格位置：%s' % start_cell)

    col_letter = col_num_to_letter(col)

    # 起始单元格设为图片大小，让图片正好填满单个单元格
    ws.column_dimensions[col_letter].width = COL_WIDTH
    ws.row_dimensions[row].height = ROW_HEIGHT

    for img_file in images:
        cell = '%s%d' % (col_letter, row)

        img = XLImage(img_file)
        # 固定 10cm x 15cm（openpyxl 单位为像素）
        img.width = IMG_W_PX
        img.height = IMG_H_PX
        ws.add_image(img, cell)

        # 从文件名解析四字段（发票号_购买方_销售方_金额）
        base = os.path.splitext(os.path.basename(img_file))[0]
        parts = base.split('_')
        # 去掉开头的全局序号
        if parts and re.match(r'^\d{4}$', parts[0]):
            parts = parts[1:]
        labels = FIELD_LABELS
        # 文件名拆分后：第1个=发票号，第2=购买方，第3=销售方，剩余拼接=金额
        no = parts[0] if len(parts) > 0 else '未知'
        buyer = parts[1] if len(parts) > 1 else '未知'
        seller = parts[2] if len(parts) > 2 else '未知'
        amount = '_'.join(parts[3:]) if len(parts) > 3 else '未知'
        # 去掉文件重名时追加的 (2) 序号
        amount = re.sub(r'\(\d+\)$', '', amount)
        vals = [no, buyer, seller, amount]

        if direction == 'h':
            # 横向：图片沿行向右，字段放图片下方（同一列、往下四行）
            text_col = col
            text_row = row + 1
            for label, val in zip(labels, vals):
                c = '%s%d' % (col_num_to_letter(text_col), text_row)
                ws[c] = '%s：%s' % (label, val)
                ws.row_dimensions[text_row].height = 18
                text_row += 1
            # 下一个图片：列 +1
            col += 1
            col_letter = col_num_to_letter(col)
            ws.column_dimensions[col_letter].width = COL_WIDTH
            ws.row_dimensions[row].height = ROW_HEIGHT
        else:
            # 纵向：图片沿列向下，字段放图片右侧（同一行、往右四列）
            text_col = col + 1
            text_row = row
            for label, val in zip(labels, vals):
                c = '%s%d' % (col_num_to_letter(text_col), text_row)
                ws[c] = '%s：%s' % (label, val)
                ws.column_dimensions[col_num_to_letter(text_col)].width = TEXT_COL_WIDTH
                text_col += 1
            # 下一个图片：行 +1
            row += 1
            col_letter = col_num_to_letter(col)
            ws.column_dimensions[col_letter].width = COL_WIDTH
            ws.row_dimensions[row].height = ROW_HEIGHT

        print('  插入 %s <- %s' % (cell, os.path.basename(img_file)))

    wb.save(xlsx_path)
    print('插入完成，共 %d 张图片，表格已保存：%s' % (len(images), xlsx_path))


# ---------------------------------------------------------------------------
# 交互流程
# ---------------------------------------------------------------------------
def existing_mode(pdf_paths):
    """现有功能：PDF 转图片 + 发票识别重命名 + 保存/插入 Excel。"""
    # 汇总输出目录：用第一个 PDF 所在目录
    base_dir = os.path.dirname(os.path.abspath(pdf_paths[0]))
    out_dir = os.path.join(base_dir, '_发票图片')
    os.makedirs(out_dir, exist_ok=True)

    # ---- 现有模式：PDF 转图片 ----
    images = []
    seq = 0
    for p in pdf_paths:
        try:
            imgs, seq = pdf_to_images(p, out_dir, start_index=seq)
            images.extend(imgs)
        except Exception as e:
            print('转换失败：%s，跳过。' % e)

    if not images:
        print('没有成功转换的图片。')
        return

    # ---- 步骤 1.5：OCR 识别并重命名 ----
    print('开始识别发票字段并重命名…')
    renamed = []
    for img in images:
        fields = extract_invoice_fields(img)
        new = rename_with_fields(img, fields)
        renamed.append(new)
        print('  重命名 -> %s' % os.path.basename(new))
    images = renamed

    # ---- 步骤 2：选择模式 ----
    mode = _ask('请选择：1 直接保存到文件夹（默认） | 2 暂存并插入表格。输入 1 或 2：').strip()
    if mode != '2':
        print('已保存到文件夹：%s' % out_dir)
        return
    if not OPENPYXL_OK:
        print('提示：未安装 openpyxl，无法写入 Excel。请运行：pip install openpyxl')
        return

    # ---- 模式 2：拖入 Excel 表格 ----
    while True:
        xlsx_paths = read_paths(
            '请拖入 Excel 表格文件（*.xlsx），然后回车；\n'
            '  输入 c 改为从剪贴板读取：')
        if not xlsx_paths:
            print('未检测到文件路径，请重新输入。')
            continue
        xlsx_path = xlsx_paths[0]
        if not (xlsx_path.lower().endswith('.xlsx') or xlsx_path.lower().endswith('.xlsm')):
            print('仅支持 .xlsx 表格，请重新输入。')
            continue
        if not os.path.isfile(xlsx_path):
            print('找不到文件：%s，请重新输入。' % xlsx_path)
            continue

        # 选择工作表
        wb = load_workbook(xlsx_path)
        sheet_names = wb.sheetnames
        print('表格内的工作表：')
        for i, name in enumerate(sheet_names, start=1):
            print('  %d. %s' % (i, name))
        while True:
            sel = _ask('请选择工作表（输入序号，回车默认 1）：').strip()
            if not sel:
                sel = '1'
            try:
                sheet_idx = int(sel)
                sheet_name = sheet_names[sheet_idx - 1]
                break
            except (ValueError, IndexError):
                print('请输入有效的序号（1-%d）。' % len(sheet_names))
        print('已选择工作表：%s' % sheet_name)

        # 输入起始单元格
        while True:
            start_cell = _ask('请输入起始单元格位置（如 C5，回车默认 A1）：').strip().upper()
            if not start_cell:
                start_cell = 'A1'
            if parse_cell(start_cell):
                break
            print('无效的单元格位置，请输入类似 C5 的坐标。')

        # 选择插入方向
        while True:
            d = _ask('请选择插入方向：1 纵向（沿列向下，默认） | 2 横向（沿行向右）：').strip()
            if d in ('', '1'):
                direction = 'v'
                break
            if d == '2':
                direction = 'h'
                break
            print('请输入 1 或 2。')

        try:
            images_into_excel(xlsx_path, images, sheet_name=sheet_name,
                              start_cell=start_cell, direction=direction)
            break
        except Exception as e:
            print('插入失败：%s，请重新输入。' % e)


# ---------------------------------------------------------------------------
def main():
    init_console_color()
    print('=' * 60)
    print('PDF 工具：现有发票图片识别 / 发票明细转表格')
    print('=' * 60)
    print('提示：把 PDF 文件拖入本窗口（可多个），然后按回车。')
    print()

    while True:
        # ---- 先选功能模式 ----
        while True:
            top_mode = _ask('请选择功能：1 收款组 PDF 转图片+发票识别 | 2 付款组 发票明细识别并转 Excel：').strip()
            if top_mode in ('1', '2'):
                break
            print('请输入 1 或 2。')

        # ---- 付款组：先选发票类型，再拖入 PDF ----
        inv_type = None
        if top_mode == '2':
            while True:
                inv_type = _ask('请选择发票类型：1 canexs | 2 精准 | 3 创时亚马逊卡派 | 4 创时卡派 | 5 创时清关费：').strip()
                if inv_type in ('1', '2', '3', '4', '5'):
                    break
                print('请输入 1 或 2。')

        # ---- 再拖入 PDF ----
        # 经典终端多选拖入只插第一个路径，可输入 'c' 从剪贴板读取全部路径
        # （在资源管理器多选文件后 Ctrl+C，剪贴板保存全部路径，每行一个）。
        while True:
            pdf_paths = read_paths(
                '请拖入 PDF 文件（可多个），然后回车；\n'
                '  输入 c 改为从剪贴板读取（资源管理器多选文件后 Ctrl+C）：')
            pdf_paths = [p for p in pdf_paths
                         if p.lower().endswith('.pdf')]
            if pdf_paths:
                break
            print('未检测到有效的 PDF 路径。若输入了 c，请先在资源管理器里选中 PDF 文件按 Ctrl+C'
                  '（复制文件本身，不是复制文字），再输入 c；或直接把文件拖入窗口。')

        if top_mode == '2':
            try:
                detail_mode(pdf_paths, inv_type)
            except Exception as e:
                print('发票明细识别失败：%s' % e)
        else:
            existing_mode(pdf_paths)

        # ---- 是否继续 ----
        again = _ask('是否继续处理其他 PDF？(y/n，回车默认 n)：').strip().lower()
        if again not in ('y', 'yes'):
            break

    print('已完成，谢谢使用。')


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('\n用户取消操作。')
    sys.exit(0)
